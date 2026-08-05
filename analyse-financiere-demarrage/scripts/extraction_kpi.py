"""
Extraction et recalcul des indicateurs financiers du bilan de démarrage.

Le classeur `bilan-demarrage.xlsx` contient 29 feuilles de travail (projections,
inventaire, montage financier, outils d'exploitation, états financiers...).
Ce script en extrait les postes utiles, RECALCULE les ratios de façon
indépendante (plutôt que de recopier les cellules de ratios du classeur), puis
écrit des fichiers CSV propres dans `../data/`.

Recalculer plutôt que recopier est volontaire : c'est le contrôle croisé qui
permet d'affirmer que les chiffres cités dans le README sont justes.

Usage :
    python extraction_kpi.py
"""

from pathlib import Path
import warnings

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")  # images WMF et en-têtes non lisibles par openpyxl

RACINE = Path(__file__).resolve().parent.parent
CLASSEUR = RACINE / "bilan-demarrage.xlsx"
DOSSIER_DONNEES = RACINE / "data"

# Valeurs d'erreur Excel présentes dans les zones de scénario non utilisées
# du classeur : elles ne doivent jamais être converties en nombre.
ERREURS_EXCEL = {"#VALUE!", "#NUM!", "#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!"}


def num(feuille, coord):
    """Renvoie la valeur numérique d'une cellule, ou None si vide ou en erreur."""
    valeur = feuille[coord].value
    if valeur is None or (isinstance(valeur, str) and valeur.strip() in ERREURS_EXCEL):
        return None
    try:
        return float(valeur)
    except (TypeError, ValueError):
        return None


def txt(feuille, coord):
    """Renvoie le libellé d'une cellule, nettoyé."""
    valeur = feuille[coord].value
    return str(valeur).strip() if valeur is not None else ""


def ratio(numerateur, denominateur):
    """Division protégée : évite la division par zéro sur un dénominateur vide."""
    if not denominateur:
        return None
    return numerateur / denominateur


# --------------------------------------------------------------------------
# 1. États des résultats — trimestres 1 et 2
# --------------------------------------------------------------------------
def extraire_resultats(wb):
    """La feuille du TRIM 2 porte les deux trimestres : B = T1, D = T2, F = cumul."""
    ws = wb["S6 États des résultats TRIM 2"]

    postes = {
        "Ventes nettes": 9,
        "Achats": 11,
        "Frais de transport sur achats": 12,
        "Stocks de début": 13,
        "Stocks de fin": 14,
        "Coût des marchandises vendues": 15,
        "Bénéfice brut": 16,
    }
    # Frais d'exploitation détaillés : lignes 20 à 40
    for ligne in range(20, 41):
        libelle = txt(ws, f"A{ligne}")
        if libelle:
            postes[libelle] = ligne
    postes["Total des frais d'exploitation"] = 41
    postes["Résultat net"] = 42

    lignes = []
    for libelle, ligne in postes.items():
        lignes.append(
            {
                "poste": libelle,
                "trim_1": num(ws, f"B{ligne}"),
                "trim_2": num(ws, f"D{ligne}"),
                "cumul_semestre": num(ws, f"F{ligne}"),
            }
        )

    df = pd.DataFrame(lignes)

    # Contrôle croisé : le cumul doit être la somme des deux trimestres.
    controle = df.dropna(subset=["trim_1", "trim_2", "cumul_semestre"]).copy()
    controle["ecart"] = (
        controle["trim_1"] + controle["trim_2"] - controle["cumul_semestre"]
    ).abs()
    ecart_max = controle["ecart"].max()
    print(f"  [contrôle] écart max cumul vs T1+T2 : {ecart_max:.2f} $")

    return df


# --------------------------------------------------------------------------
# 2. Bilans — trimestres 1 et 2
# --------------------------------------------------------------------------
def extraire_bilans(wb):
    """
    Feuille `S8 Bilan TRIM2`. Les postes de détail sont en B (T1) et E (T2),
    les sous-totaux en C (T1) et F (T2) — d'où la colonne explicite par poste.
    """
    ws = wb["S8 Bilan TRIM2 "]

    # (libellé, ligne, colonne T1, colonne T2, catégorie)
    postes = [
        ("Banque", 9, "B", "E", "Actif à court terme"),
        ("Comptes clients", 10, "B", "E", "Actif à court terme"),
        ("Charges payées d'avance", 11, "B", "E", "Actif à court terme"),
        ("Stocks de marchandises", 12, "B", "E", "Actif à court terme"),
        ("Total de l'actif à court terme", 13, "C", "F", "Sous-total"),
        ("Total des immobilisations", 34, "C", "F", "Sous-total"),
        ("Total de l'actif", 35, "C", "F", "Total"),
        ("Comptes fournisseurs", 39, "B", "F", "Passif à court terme"),
        ("Salaires à payer", 40, "B", "F", "Passif à court terme"),
        ("Vacances à payer", 41, "B", "F", "Passif à court terme"),
        ("D.A.S. à payer", 42, "B", "F", "Passif à court terme"),
        ("Taxes à payer", 43, "B", "F", "Passif à court terme"),
        ("Capital de risque", 44, "B", "F", "Passif à court terme"),
        ("Total du passif à court terme", 45, "B", "F", "Sous-total"),
        ("Emprunt hypothécaire", 48, "B", "F", "Passif à long terme"),
        ("Total du passif", 50, "B", "F", "Total"),
        ("Capital-actions", 53, "B", "F", "Capitaux propres"),
        ("Subventions", 54, "B", "F", "Capitaux propres"),
        ("Résultat net cumulé", 55, "B", "F", "Capitaux propres"),
        ("Total des capitaux propres", 56, "C", "F", "Total"),
    ]

    lignes = []
    for libelle, ligne, col_t1, col_t2, categorie in postes:
        lignes.append(
            {
                "poste": libelle,
                "categorie": categorie,
                "trim_1": num(ws, f"{col_t1}{ligne}"),
                "trim_2": num(ws, f"{col_t2}{ligne}"),
            }
        )

    df = pd.DataFrame(lignes)

    # Analyse verticale (% du total de l'actif) et horizontale (variation T1 → T2),
    # recalculées ici plutôt que reprises des colonnes D/G/H du classeur.
    actif_t1 = df.loc[df["poste"] == "Total de l'actif", "trim_1"].iloc[0]
    actif_t2 = df.loc[df["poste"] == "Total de l'actif", "trim_2"].iloc[0]
    df["av_trim_1"] = df["trim_1"] / actif_t1
    df["av_trim_2"] = df["trim_2"] / actif_t2
    df["ah_t1_t2"] = df.apply(
        lambda r: ratio(r["trim_2"] - r["trim_1"], abs(r["trim_1"]))
        if r["trim_1"] not in (None, 0)
        else None,
        axis=1,
    )

    # Contrôle croisé : l'équation comptable doit tenir aux deux trimestres.
    passif_t1 = df.loc[df["poste"] == "Total du passif", "trim_1"].iloc[0]
    passif_t2 = df.loc[df["poste"] == "Total du passif", "trim_2"].iloc[0]
    capitaux_t1 = df.loc[df["poste"] == "Total des capitaux propres", "trim_1"].iloc[0]
    capitaux_t2 = df.loc[df["poste"] == "Total des capitaux propres", "trim_2"].iloc[0]
    print(
        f"  [contrôle] Actif - (Passif + Capitaux) : "
        f"T1 {actif_t1 - passif_t1 - capitaux_t1:+.2f} $ · "
        f"T2 {actif_t2 - passif_t2 - capitaux_t2:+.2f} $"
    )

    return df


# --------------------------------------------------------------------------
# 3. Ratios financiers — recalculés à partir des états extraits
# --------------------------------------------------------------------------
def calculer_ratios(resultats, bilans):
    def poste_res(nom, colonne):
        return resultats.loc[resultats["poste"] == nom, colonne].iloc[0]

    def poste_bil(nom, colonne):
        return bilans.loc[bilans["poste"] == nom, colonne].iloc[0]

    lignes = []
    for trim, col in (("Trim 1", "trim_1"), ("Trim 2", "trim_2")):
        ventes = poste_res("Ventes nettes", col)
        benefice_brut = poste_res("Bénéfice brut", col)
        resultat = poste_res("Résultat net", col)
        actif = poste_bil("Total de l'actif", col)
        act = poste_bil("Total de l'actif à court terme", col)
        pct = poste_bil("Total du passif à court terme", col)
        passif = poste_bil("Total du passif", col)
        capitaux = poste_bil("Total des capitaux propres", col)
        resultat_cumule = poste_bil("Résultat net cumulé", col)

        lignes.append(
            {
                "trimestre": trim,
                "marge_brute": ratio(benefice_brut, ventes),
                "marge_nette": ratio(resultat, ventes),
                "rotation_actif": ratio(ventes, actif),
                "levier_financier": ratio(actif, capitaux),
                "roe_cumule": ratio(resultat_cumule, capitaux),
                "liquidite_generale": ratio(act, pct),
                "fonds_roulement": act - pct,
                "endettement": ratio(passif, actif),
            }
        )

    df = pd.DataFrame(lignes)

    # Contrôle croisé DuPont : marge nette × rotation × levier = rendement du
    # trimestre sur les capitaux propres.
    df["dupont_trimestriel"] = (
        df["marge_nette"] * df["rotation_actif"] * df["levier_financier"]
    )
    return df


# --------------------------------------------------------------------------
# 4. Seuil de rentabilité
# --------------------------------------------------------------------------
def extraire_seuil(wb, resultats):
    ws = wb["Fx var et Fx fixe -Simul"]

    ventes_semestre = num(ws, "H5")
    frais_variables = num(ws, "H6")
    marge_cout_variable = num(ws, "H7")
    frais_fixes_semestre = num(ws, "H8")
    frais_fixes_annuels = num(ws, "H11")

    taux_mcv = ratio(marge_cout_variable, ventes_semestre)
    seuil_recalcule = ratio(frais_fixes_annuels, taux_mcv)
    seuil_classeur = num(ws, "G20")

    print(
        f"  [contrôle] seuil recalculé {seuil_recalcule:,.0f} $ vs "
        f"classeur {seuil_classeur:,.0f} $ "
        f"(écart {abs(seuil_recalcule - seuil_classeur):,.2f} $)"
    )

    scenarios = [
        ("Seuil de rentabilité", num(ws, "G20")),
        ("Prévisions année 1", num(ws, "H20")),
        ("Prévisions année 2", num(ws, "I20")),
        ("Prévisions année 3", num(ws, "J20")),
        ("Bénéfice net visé de 500 000 $", num(ws, "K20")),
        ("Semestre 1 réel annualisé", num(ws, "L20")),
    ]
    df_scenarios = pd.DataFrame(scenarios, columns=["scenario", "ventes_requises"])
    df_scenarios["marge_securite"] = df_scenarios["ventes_requises"] - seuil_recalcule
    df_scenarios["taux_marge_securite"] = (
        df_scenarios["marge_securite"] / df_scenarios["ventes_requises"]
    )

    # --- Analyse de sensibilité (hors classeur) --------------------------
    # Les frais fixes annuels sont obtenus en doublant ceux du semestre 1. Or le
    # semestre 1 contient une campagne de lancement non récurrente : la publicité
    # passe de 276 244 $ au T1 à 9 000 $ au T2, soit un régime de croisière de
    # 3 000 $/mois. L'annualisation reconduit donc DEUX FOIS une dépense unique,
    # ce qui gonfle le seuil de rentabilité.
    pub_t1 = resultats.loc[resultats["poste"] == "Publicité", "trim_1"].iloc[0]
    pub_t2 = resultats.loc[resultats["poste"] == "Publicité", "trim_2"].iloc[0]
    lancement = pub_t1 - pub_t2
    fixes_regime = frais_fixes_annuels - 2 * lancement
    seuil_regime = ratio(fixes_regime, taux_mcv)
    print(
        f"  [sensibilité] campagne de lancement non récurrente : {lancement:,.0f} $ "
        f"-> frais fixes en régime {fixes_regime:,.0f} $, "
        f"seuil {seuil_regime:,.0f} $ (au lieu de {seuil_recalcule:,.0f} $)"
    )

    synthese = {
        "ventes_semestre": ventes_semestre,
        "frais_variables": frais_variables,
        "taux_frais_variables": ratio(frais_variables, ventes_semestre),
        "marge_cout_variable": marge_cout_variable,
        "taux_marge_cout_variable": taux_mcv,
        "frais_fixes_semestre": frais_fixes_semestre,
        "frais_fixes_annuels": frais_fixes_annuels,
        "seuil_rentabilite": seuil_recalcule,
        "campagne_lancement": lancement,
        "frais_fixes_regime": fixes_regime,
        "seuil_regime": seuil_regime,
    }
    return synthese, df_scenarios


# --------------------------------------------------------------------------
# 5. Rentabilité par produit
# --------------------------------------------------------------------------
def extraire_rentabilite_produits(wb):
    ws = wb["S7 - Rentabilité par prod. ETUD"]

    lignes = []
    for ligne in range(20, 33):  # P1 à P13
        lignes.append(
            {
                "produit": txt(ws, f"A{ligne}"),
                "prix_vente": num(ws, f"B{ligne}"),
                "cout_achat": num(ws, f"C{ligne}"),
                "transport_achat": num(ws, f"E{ligne}"),
                "commission": num(ws, f"G{ligne}"),
                "avantages_sociaux": num(ws, f"H{ligne}"),
                "frais_livraison": num(ws, f"J{ligne}") or 0.0,
                "honoraires": num(ws, f"L{ligne}"),
                "total_couts_directs": num(ws, f"N{ligne}"),
                "marge_unitaire": num(ws, f"P{ligne}"),
                "marge_volume_semestre": num(ws, f"R{ligne}"),
                "marge_volume_croisiere": num(ws, f"T{ligne}"),
            }
        )

    df = pd.DataFrame(lignes)

    # Recalcul indépendant du taux de marge sur coûts directs.
    df["taux_marge"] = df["marge_unitaire"] / df["prix_vente"]
    df["part_marge_semestre"] = (
        df["marge_volume_semestre"] / df["marge_volume_semestre"].sum()
    )
    df["part_marge_croisiere"] = (
        df["marge_volume_croisiere"] / df["marge_volume_croisiere"].sum()
    )

    ecart = (df["prix_vente"] - df["total_couts_directs"] - df["marge_unitaire"]).abs().max()
    print(f"  [contrôle] écart max prix - coûts - marge : {ecart:.4f} $")

    return df


# --------------------------------------------------------------------------
# 6. Prévisions de ventes sur 3 ans
# --------------------------------------------------------------------------
def extraire_previsions(wb):
    """Ligne 37 de `S1 - projections` : ventes totales en $ par mois."""
    ws = wb["S1 - projections"]

    plages = {
        1: ("C", "N", "O"),   # année 1 : mois en C..N, total annuel en O
        2: ("S", "AD", "AE"),
        3: ("AI", "AT", "AU"),
    }

    def colonnes(debut, fin):
        i_debut = openpyxl.utils.column_index_from_string(debut)
        i_fin = openpyxl.utils.column_index_from_string(fin)
        return [
            openpyxl.utils.get_column_letter(i) for i in range(i_debut, i_fin + 1)
        ]

    lignes = []
    for annee, (debut, fin, col_total) in plages.items():
        cols = colonnes(debut, fin)
        total_classeur = num(ws, f"{col_total}37")
        somme = 0.0
        for mois, col in enumerate(cols, start=1):
            valeur = num(ws, f"{col}37") or 0.0
            somme += valeur
            lignes.append(
                {
                    "annee": annee,
                    "mois": mois,
                    "mois_absolu": (annee - 1) * 12 + mois,
                    "trimestre": (mois - 1) // 3 + 1,
                    "ventes": valeur,
                }
            )
        print(
            f"  [contrôle] année {annee} : somme mensuelle {somme:,.0f} $ vs "
            f"total classeur {total_classeur:,.0f} $ "
            f"(écart {abs(somme - total_classeur):,.2f} $)"
        )

    return pd.DataFrame(lignes)


# --------------------------------------------------------------------------
# 7. Structure de financement au démarrage
# --------------------------------------------------------------------------
def extraire_montage(wb):
    ws = wb["S4 - Montage financier"]

    actifs = [
        ("Taxes à recevoir", "C2"),
        ("Stock de marchandises", "C3"),
        ("Charges payées d'avance", "C4"),
        ("Frais d'incorporation", "C5"),
        ("Améliorations locatives", "C6"),
        ("Caisses enregistreuses", "C7"),
        ("Ameublement", "C8"),
        ("Équipement informatique", "C9"),
        ("Équipement de manutention", "C10"),
        ("Équipement de cuisine", "C11"),
    ]
    passifs = [
        ("Marge de crédit", "D2"),
        ("Comptes fournisseurs", "D3"),
        ("Emprunt hypothécaire", "D4"),
        ("Capital de risque", "D5"),
        ("Capital-actions", "D10"),
        ("Subvention", "D11"),
    ]

    lignes = [
        {"volet": "Actif", "poste": nom, "montant": num(ws, coord)}
        for nom, coord in actifs
    ] + [
        {"volet": "Passif et capitaux propres", "poste": nom, "montant": num(ws, coord)}
        for nom, coord in passifs
    ]
    df = pd.DataFrame(lignes)

    total_actif = df.loc[df["volet"] == "Actif", "montant"].sum()
    total_passif = df.loc[df["volet"] != "Actif", "montant"].sum()
    print(
        f"  [contrôle] montage (option location) : actif {total_actif:,.0f} $ vs "
        f"passif+capitaux {total_passif:,.0f} $ "
        f"(écart {abs(total_actif - total_passif):,.2f} $)"
    )
    return df


# --------------------------------------------------------------------------
# Programme principal
# --------------------------------------------------------------------------
def main():
    DOSSIER_DONNEES.mkdir(exist_ok=True)
    print(f"Lecture de {CLASSEUR.name} ...")
    wb = openpyxl.load_workbook(CLASSEUR, data_only=True)
    print(f"  {len(wb.sheetnames)} feuilles détectées\n")

    print("Extraction des états des résultats")
    resultats = extraire_resultats(wb)

    print("Extraction des bilans")
    bilans = extraire_bilans(wb)

    print("Calcul des ratios")
    ratios = calculer_ratios(resultats, bilans)

    print("Extraction du seuil de rentabilité")
    seuil, scenarios = extraire_seuil(wb, resultats)

    print("Extraction de la rentabilité par produit")
    produits = extraire_rentabilite_produits(wb)

    print("Extraction des prévisions de ventes")
    previsions = extraire_previsions(wb)

    print("Extraction du montage financier")
    montage = extraire_montage(wb)

    sorties = {
        "etats_resultats.csv": resultats,
        "bilans.csv": bilans,
        "ratios.csv": ratios,
        "seuil_scenarios.csv": scenarios,
        "rentabilite_produits.csv": produits,
        "previsions_ventes.csv": previsions,
        "montage_financier.csv": montage,
        "seuil_synthese.csv": pd.DataFrame([seuil]),
    }

    print("\nÉcriture des fichiers CSV")
    for nom, df in sorties.items():
        chemin = DOSSIER_DONNEES / nom
        df.to_csv(chemin, index=False, encoding="utf-8-sig")
        print(f"  {nom:<28} {len(df):>3} lignes")

    print("\nTerminé.")


if __name__ == "__main__":
    main()
