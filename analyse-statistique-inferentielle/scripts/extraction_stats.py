"""
Extraction et recalcul des analyses statistiques du classeur `analyse-statistique.xlsx`.

Le classeur contient trois échantillons bruts (225 employés d'Abus inc.,
150 manœuvres beaucerons, 2 330 pièces de Kansas Vamal) et onze feuilles
d'analyse : distribution, corrélation, régression, intervalle de confiance et
quatre tests d'hypothèses.

Ce script ne recopie AUCUN résultat du classeur. Il repart des données brutes,
recalcule chaque mesure et chaque test avec ses propres fonctions de loi de
probabilité (aucune dépendance à scipy), puis vérifie que le résultat obtenu
coïncide avec celui du classeur. Les 36 contrôles doivent ressortir à zéro
d'écart : c'est ce qui atteste que les chiffres publiés sont justes.

Les fonctions de loi implémentées ici (Phi, invPhi, chi2_cdf, chi2_inv)
reproduisent NORM.DIST, NORM.INV, CHIDIST et KHIDEUX.INVERSE.DROITE d'Excel.

Usage :
    python extraction_stats.py
"""

from pathlib import Path
import math
import sys
import warnings

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")  # en-têtes et objets non lisibles par openpyxl

RACINE = Path(__file__).resolve().parent.parent
CLASSEUR = RACINE / "analyse-statistique.xlsx"
DOSSIER_DONNEES = RACINE / "data"

FEUILLE_ABUS = "Abus inc."
FEUILLE_BEAUCERONS = "Manoeuvres beaucerons"
FEUILLE_KANSAS = "Kansas Vamal"

controles = []



def controler(libelle, recalcule, classeur, tolerance=5e-4):
    """Compare une valeur recalculée à la valeur inscrite dans le classeur."""
    ecart = None if classeur is None else abs(recalcule - classeur)
    controles.append(
        {
            "controle": libelle,
            "recalcule": recalcule,
            "classeur": classeur,
            "ecart": ecart,
            "statut": "ok" if (ecart is not None and ecart <= tolerance) else "à expliquer",
        }
    )


# =========================================================================
# Lois de probabilité (équivalents Excel, sans scipy)
# =========================================================================
def Phi(z):
    """Fonction de répartition de la loi normale centrée réduite."""
    return 0.5 * (1 + math.erf(z / math.sqrt(2)))


def invPhi(p):
    """Quantile de la loi normale centrée réduite, par dichotomie."""
    bas, haut = -12.0, 12.0
    for _ in range(200):
        milieu = (bas + haut) / 2
        if Phi(milieu) < p:
            bas = milieu
        else:
            haut = milieu
    return (bas + haut) / 2


def gamma_reg(s, x):
    """Fonction gamma incomplète régularisée P(s, x), série puis fraction continue."""
    if x <= 0:
        return 0.0
    if x < s + 1:
        terme = 1.0 / s
        total = terme
        for n in range(1, 2000):
            terme *= x / (s + n)
            total += terme
            if abs(terme) < 1e-16 * abs(total):
                break
        return total * math.exp(-x + s * math.log(x) - math.lgamma(s))
    minuscule = 1e-300
    b, c, d = x + 1 - s, 1 / minuscule, 1 / (x + 1 - s)
    h = d
    for i in range(1, 2000):
        an = -i * (i - s)
        b += 2
        d = an * d + b
        d = minuscule if abs(d) < minuscule else d
        c = b + an / c
        c = minuscule if abs(c) < minuscule else c
        d = 1 / d
        delta = d * c
        h *= delta
        if abs(delta - 1) < 1e-15:
            break
    return 1 - math.exp(-x + s * math.log(x) - math.lgamma(s)) * h


def chi2_cdf(x, dl):
    return gamma_reg(dl / 2, x / 2)


def chi2_inv(p, dl):
    """Quantile du khi-deux, par dichotomie."""
    bas, haut = 0.0, 2000.0
    for _ in range(300):
        milieu = (bas + haut) / 2
        if chi2_cdf(milieu, dl) < p:
            bas = milieu
        else:
            haut = milieu
    return (bas + haut) / 2


# =========================================================================
# 1. Lecture des trois échantillons bruts
# =========================================================================
def lire_echantillons(wb):
    ws = wb[FEUILLE_ABUS]
    lignes = [
        r[:10] for r in ws.iter_rows(min_row=5, values_only=True) if r[0] is not None
    ]
    abus = pd.DataFrame(
        lignes,
        columns=[
            "sexe",
            "anciennete",
            "jours_absence",
            "aucun_acte",
            "attitude_negative",
            "refus_autorite",
            "bris_intentionnels",
            "agression",
            "harcelement",
            "total_actes",
        ],
    )
    abus["rebelle"] = abus["total_actes"] > 0

    ws = wb[FEUILLE_BEAUCERONS]
    lignes = [
        r[:2] for r in ws.iter_rows(min_row=4, values_only=True) if r[0] is not None
    ]
    beaucerons = pd.DataFrame(lignes, columns=["salaire", "experience"])

    ws = wb[FEUILLE_KANSAS]
    pieces = [
        r[0] for r in ws.iter_rows(min_row=4, values_only=True) if r[0] is not None
    ]

    controler("Effectif de l'échantillon Abus inc.", len(abus), 225, 0)
    controler("Effectif de l'échantillon beaucerons", len(beaucerons), 150, 0)
    controler("Effectif de l'échantillon Kansas Vamal", len(pieces), 2330, 0)
    return abus, beaucerons, pieces


# =========================================================================
# 2. Distribution et mesures de l'ancienneté
# =========================================================================
def grouper(serie, bornes, gabarit):
    """Effectifs, fréquences et fréquences cumulées pour une série de bornes."""
    classes, cumul = [], 0
    for i in range(len(bornes) - 1):
        bas, haut = bornes[i], bornes[i + 1]
        effectif = int(((serie >= bas) & (serie < haut)).sum())
        cumul += effectif
        classes.append(
            {
                "classe": gabarit.format(bas=bas, haut=haut),
                "borne_inf": bas,
                "borne_sup": haut,
                "effectif": effectif,
                "frequence": effectif / len(serie),
                "frequence_cumulee": cumul / len(serie),
            }
        )
    return pd.DataFrame(classes)


def analyser_anciennete(abus, wb):
    ws = wb["Ancienneté – SD "]
    anc = abus["anciennete"]

    # Le tableau croisé dynamique du classeur groupe par pas de 3 années à
    # partir de la valeur minimale observée (1,5966), et non à partir d'une
    # borne ronde. Ce sont ces classes qui produisent les effectifs du classeur.
    depart = anc.min()
    bornes_classeur = [depart + 3 * i for i in range(9)]
    distribution = grouper(anc, bornes_classeur, "[{bas:.2f} - {haut:.2f}[")
    controler(
        "Somme des effectifs par classe d'ancienneté",
        int(distribution["effectif"].sum()),
        len(anc),
        0,
    )

    # Le tableau de fréquences présenté occupe les lignes 21 à 28, colonne D.
    effectifs_classeur = [ws[f"D{ligne}"].value for ligne in range(21, 29)]
    controler(
        "Écart total entre les effectifs par classe et ceux du classeur",
        float(
            sum(abs(a - b) for a, b in zip(distribution["effectif"], effectifs_classeur))
        ),
        0.0,
        0,
    )

    # Les libellés doivent nommer ces mêmes bornes. Sur des bornes rondes
    # ([1-5[, [5-8[...), 32 employés se retrouveraient dans une autre classe.
    libelles_classeur = [ws[f"C{ligne}"].value for ligne in range(21, 29)]
    controler(
        "Libellés de classes ne nommant pas les bornes réelles",
        float(
            sum(
                1
                for attendu, lu in zip(distribution["classe"], libelles_classeur)
                if attendu.replace(".", ",") != str(lu).strip()
            )
        ),
        0.0,
        0,
    )

    # Mêmes données sur des bornes rondes, à titre de comparaison.
    distribution_ronde = grouper(anc, [1, 4, 7, 10, 13, 16, 19, 22, 25], "[{bas}-{haut}[")

    mesures = {
        "n": len(anc),
        "moyenne": anc.mean(),
        "mediane": anc.median(),
        "ecart_type": anc.std(ddof=1),
        "coefficient_variation": anc.std(ddof=1) / anc.mean(),
        "minimum": anc.min(),
        "maximum": anc.max(),
        "q1": anc.quantile(0.25),
        "q3": anc.quantile(0.75),
    }
    controler("Ancienneté moyenne", mesures["moyenne"], ws["D55"].value)
    controler("Écart-type de l'ancienneté", mesures["ecart_type"], ws["D56"].value)
    controler("Coefficient de variation", mesures["coefficient_variation"], ws["D57"].value)
    controler("Médiane de l'ancienneté", mesures["mediane"], ws["D58"].value)

    return distribution, distribution_ronde, pd.DataFrame([mesures])


# =========================================================================
# 3. Répartition des actes de rébellion
# =========================================================================
def analyser_rebellion(abus, wb):
    ws = wb[" Rébellion – SD"]
    colonnes = [
        ("1 - Attitude négative", "attitude_negative"),
        ("2 - Refus de l'autorité", "refus_autorite"),
        ("3 - Bris intentionnels", "bris_intentionnels"),
        ("4 - Agression", "agression"),
        ("5 - Harcèlement", "harcelement"),
    ]
    total = sum(int(abus[c].sum()) for _, c in colonnes)
    lignes = []
    for libelle, colonne in colonnes:
        effectif = int(abus[colonne].sum())
        lignes.append(
            {
                "acte": libelle,
                "nombre_actes": effectif,
                "part_des_actes": effectif / total,
            }
        )
    actes = pd.DataFrame(lignes)
    controler("Nombre total d'actes de rébellion", total, ws["C14"].value, 0)
    controler(
        "Actes de type « attitude négative »",
        int(abus["attitude_negative"].sum()),
        ws["C9"].value,
        0,
    )

    profil = pd.DataFrame(
        [
            {
                "employes": len(abus),
                "employes_rebelles": int(abus["rebelle"].sum()),
                "part_rebelles": abus["rebelle"].mean(),
                "actes_totaux": total,
                "actes_par_rebelle": total / int(abus["rebelle"].sum()),
                "actes_max_un_employe": int(abus["total_actes"].max()),
                "corr_anciennete_actes": abus["anciennete"].corr(abus["total_actes"]),
                "corr_absence_actes": abus["jours_absence"].corr(abus["total_actes"]),
            }
        ]
    )
    return actes, profil


# =========================================================================
# 4. Corrélation et régression : ancienneté et absences
# =========================================================================
def analyser_regression(abus, wb):
    ws = wb[" Ancienneté – Absence"]
    x, y = abus["anciennete"], abus["jours_absence"]
    n = len(x)

    r = x.corr(y)
    pente = r * y.std(ddof=1) / x.std(ddof=1)
    ordonnee = y.mean() - pente * x.mean()

    controler("Coefficient de corrélation", r, ws["E3"].value)
    controler("Pente de la droite de régression", pente, ws["E29"].value)
    controler("Coefficient de détermination", r ** 2, ws["E33"].value, 5e-4)

    def predire(valeur):
        return ordonnee + pente * valeur

    pente_inverse = r * x.std(ddof=1) / y.std(ddof=1)

    parametres = pd.DataFrame(
        [
            {
                "n": n,
                "r": r,
                "r_carre": r ** 2,
                "pente": pente,
                "ordonnee_origine": ordonnee,
                "anciennete_min": x.min(),
                "anciennete_max": x.max(),
                "prediction_20_55": predire(20.55),
                "prediction_42": predire(42),
                # Pente de la régression de x sur y : la confondre avec la
                # pente ci-dessus est l'erreur classique de ce calcul. Elle
                # donnerait 13,46 jours au lieu de 8,45 à 20,55 années.
                "pente_inverse_x_sur_y": pente_inverse,
                "prediction_20_55_pente_inverse": (
                    x.mean() + pente_inverse * (20.55 - y.mean())
                ),
            }
        ]
    )
    controler("Prédiction à 20,55 années", predire(20.55), ws["G24"].value)
    controler("Prédiction à 42 années", predire(42), ws["G26"].value)
    return parametres


def qualite_absences(abus, regression):
    """Isole les enregistrements à zéro jour d'absence et mesure leur effet.

    Seize employés déclarent exactement zéro jour d'absence, à des anciennetés
    réparties sur tout l'intervalle. Leurs résidus sont tous très en dessous de
    la droite alors que ceux des autres employés tiennent dans une bande
    étroite : la structure est celle d'une non-réponse codée 0, pas celle d'une
    assiduité parfaite.
    """
    reg = regression.iloc[0]
    x, y = abus["anciennete"], abus["jours_absence"]
    residus = y - (reg["ordonnee_origine"] + reg["pente"] * x)

    sans_zero = abus[abus["jours_absence"] > 0]
    r_sans = sans_zero["anciennete"].corr(sans_zero["jours_absence"])
    pente_sans = (
        r_sans
        * sans_zero["jours_absence"].std(ddof=1)
        / sans_zero["anciennete"].std(ddof=1)
    )
    ordonnee_sans = (
        sans_zero["jours_absence"].mean() - pente_sans * sans_zero["anciennete"].mean()
    )

    zeros = abus["jours_absence"] == 0
    return pd.DataFrame(
        [
            {
                "groupe": "Échantillon complet",
                "n": len(abus),
                "r": reg["r"],
                "r_carre": reg["r_carre"],
                "pente": reg["pente"],
                "ordonnee_origine": reg["ordonnee_origine"],
                "prediction_20_55": reg["prediction_20_55"],
                "residu_min": residus.min(),
                "residu_max": residus.max(),
                "ecart_type_residus": residus.std(ddof=1),
            },
            {
                "groupe": "Sans les employés à zéro jour d'absence",
                "n": len(sans_zero),
                "r": r_sans,
                "r_carre": r_sans ** 2,
                "pente": pente_sans,
                "ordonnee_origine": ordonnee_sans,
                "prediction_20_55": ordonnee_sans + pente_sans * 20.55,
                "residu_min": residus[~zeros].min(),
                "residu_max": residus[~zeros].max(),
                "ecart_type_residus": residus[~zeros].std(ddof=1),
            },
            {
                "groupe": "Employés à zéro jour d'absence seulement",
                "n": int(zeros.sum()),
                "r": None,
                "r_carre": None,
                "pente": None,
                "ordonnee_origine": None,
                "prediction_20_55": None,
                "residu_min": residus[zeros].min(),
                "residu_max": residus[zeros].max(),
                "ecart_type_residus": residus[zeros].std(ddof=1),
            },
        ]
    )


# =========================================================================
# 5. Test du khi-deux d'indépendance : sexe et rébellion
# =========================================================================
def test_independance(abus, wb):
    ws = wb["Sexe-Rebellion"]
    seuil = ws["B44"].value

    croise = (
        pd.crosstab(abus["sexe"], abus["rebelle"])
        .rename(columns={False: "aucun_acte", True: "au_moins_un_acte"})
        .reset_index()
    )
    croise["total"] = croise["aucun_acte"] + croise["au_moins_un_acte"]
    croise["taux_rebellion"] = croise["au_moins_un_acte"] / croise["total"]

    observes = croise[["aucun_acte", "au_moins_un_acte"]].values.astype(float)
    n = observes.sum()
    marges_lignes = observes.sum(axis=1)
    marges_colonnes = observes.sum(axis=0)

    detail, khi2 = [], 0.0
    for i, sexe in enumerate(croise["sexe"]):
        for j, colonne in enumerate(["aucun_acte", "au_moins_un_acte"]):
            theorique = marges_lignes[i] * marges_colonnes[j] / n
            contribution = (observes[i, j] - theorique) ** 2 / theorique
            khi2 += contribution
            detail.append(
                {
                    "cellule": f"{sexe} · {colonne}",
                    "observe": observes[i, j],
                    "theorique": theorique,
                    "contribution": contribution,
                }
            )
    dl = 1
    critique = chi2_inv(1 - seuil, dl)
    p_valeur = 1 - chi2_cdf(khi2, dl)

    controler("Khi-deux d'indépendance calculé", khi2, ws["D41"].value)
    controler("Khi-deux critique (seuil 10 %)", critique, ws["B46"].value)
    controler(
        "Effectif « femmes sans acte » du tableau croisé",
        int(croise.loc[croise["sexe"] == "Femme", "aucun_acte"].iloc[0]),
        ws["B12"].value,
        0,
    )
    controler("Total du tableau croisé", int(n), ws["D14"].value, 0)

    resume = {
        "test": "Khi-deux d'indépendance · sexe et rébellion",
        "h0": "Aucun lien entre le sexe et l'exécution d'un acte de rébellion",
        "h1": "Il existe un lien entre le sexe et l'exécution d'un acte de rébellion",
        "seuil": seuil,
        "statistique": khi2,
        "valeur_critique": critique,
        "p_valeur": p_valeur,
        "decision": "rejet de H0" if khi2 > critique else "H0 non rejetée",
        "statistique_classeur": ws["D41"].value,
        "decision_classeur": "rejet de H0",
    }
    return croise, pd.DataFrame(detail), resume


# =========================================================================
# 6. Intervalle de confiance sur l'ancienneté moyenne
# =========================================================================
def intervalle_confiance(abus, wb):
    ws = wb["Ancienneté IC"]
    niveau = ws["D3"].value
    N = int(ws["D7"].value)
    x = abus["anciennete"]
    n, moyenne, ecart_type = len(x), x.mean(), x.std(ddof=1)

    z = invPhi(1 - (1 - niveau) / 2)
    erreur_type = ecart_type / math.sqrt(n)
    correction = math.sqrt((N - n) / (N - 1))

    lignes = []
    for libelle, marge in [
        ("Sans correction de population finie", z * erreur_type),
        ("Avec correction de population finie", z * erreur_type * correction),
    ]:
        lignes.append(
            {
                "methode": libelle,
                "niveau_confiance": niveau,
                "n": n,
                "N": N,
                "taux_sondage": n / N,
                "z": z,
                "erreur_type": erreur_type * (1 if "Sans" in libelle else correction),
                "marge_erreur": marge,
                "borne_inferieure": moyenne - marge,
                "borne_superieure": moyenne + marge,
                "largeur": 2 * marge,
            }
        )
    ic = pd.DataFrame(lignes)
    # Le taux de sondage dépasse 50 % : la marge d'erreur du classeur doit
    # porter la correction de population finie.
    controler("Marge d'erreur du classeur", z * erreur_type * correction, ws["D11"].value)
    controler(
        "Borne supérieure du classeur", moyenne + z * erreur_type * correction, ws["D13"].value
    )
    controler(
        "Facteur de correction appliqué par le classeur",
        correction,
        ws["D11"].value / (z * erreur_type),
    )
    return ic


# =========================================================================
# 7. Khi-deux d'ajustement : normalité des salaires beaucerons
# =========================================================================
def test_normalite(beaucerons, wb):
    ws = wb["Salaires beaucerons"]
    mu = ws["G4"].value
    sigma = ws["G5"].value
    seuil = ws["G27"].value
    salaires = beaucerons["salaire"]
    n = len(salaires)

    bornes = [None, 20000, 26000, 32000, 38000, 44000, 50000, None]
    lignes, khi2 = [], 0.0
    for i in range(len(bornes) - 1):
        bas, haut = bornes[i], bornes[i + 1]
        if bas is None:
            observe = int((salaires < haut).sum())
            libelle = f"moins de {haut:,} $".replace(",", " ")
            part = Phi((haut - mu) / sigma)
        elif haut is None:
            observe = int((salaires >= bas).sum())
            libelle = f"{bas:,} $ et plus".replace(",", " ")
            part = 1 - Phi((bas - mu) / sigma)
        else:
            observe = int(((salaires >= bas) & (salaires < haut)).sum())
            libelle = f"{bas:,} à {haut:,} $".replace(",", " ")
            part = Phi((haut - mu) / sigma) - Phi((bas - mu) / sigma)
        theorique = part * n
        contribution = (observe - theorique) ** 2 / theorique
        khi2 += contribution
        lignes.append(
            {
                "classe": libelle,
                "effectif_observe": observe,
                "part_theorique": part,
                "effectif_theorique": theorique,
                "contribution_khi2": contribution,
            }
        )
    distribution = pd.DataFrame(lignes)
    dl = len(lignes) - 1
    critique = chi2_inv(1 - seuil, dl)
    p_valeur = 1 - chi2_cdf(khi2, dl)

    controler("Effectifs observés par classe de salaire", distribution["effectif_observe"].sum(), n, 0)
    controler("Khi-deux d'ajustement calculé", khi2, ws["K20"].value)
    controler("Khi-deux critique d'ajustement", critique, ws["I28"].value)
    controler("Effectif théorique minimal (condition T ≥ 5)", distribution["effectif_theorique"].min(), ws["I13"].value)

    resume = {
        "test": "Khi-deux d'ajustement · normalité des salaires",
        "h0": "Le salaire des manœuvres beaucerons suit N(36 300 ; 9 500²)",
        "h1": "Le salaire des manœuvres beaucerons ne suit pas cette loi normale",
        "seuil": seuil,
        "statistique": khi2,
        "valeur_critique": critique,
        "p_valeur": p_valeur,
        "decision": "rejet de H0" if khi2 > critique else "H0 non rejetée",
        "statistique_classeur": ws["K20"].value,
        "decision_classeur": "H0 non rejetée",
    }

    ecarts = pd.DataFrame(
        [
            {
                "n": n,
                "moyenne_observee": salaires.mean(),
                "ecart_type_observe": salaires.std(ddof=1),
                "moyenne_theorique": mu,
                "ecart_type_theorique": sigma,
                "q3_theorique": mu + invPhi(0.75) * sigma,
                "q3_observe": salaires.quantile(0.75),
            }
        ]
    )
    return distribution, ecarts, resume


# =========================================================================
# 8. Test sur une moyenne : salaire des manœuvres expérimentés
# =========================================================================
def test_moyenne(beaucerons, wb):
    ws = wb["Salaires Beaucerons Expérience "]
    valeur_testee = ws["E4"].value
    seuil = ws["K2"].value
    experience_min = 11

    groupe = beaucerons[beaucerons["experience"] >= experience_min]["salaire"]
    n, moyenne, ecart_type = len(groupe), groupe.mean(), groupe.std(ddof=1)
    erreur_type = ecart_type / math.sqrt(n)
    z = (moyenne - valeur_testee) / erreur_type
    critique = valeur_testee - invPhi(1 - seuil) * erreur_type
    p_valeur = Phi(z)

    controler("Effectif du sous-échantillon expérimenté", n, ws["E2"].value, 0)
    controler("Salaire moyen du sous-échantillon", moyenne, ws["G2"].value)
    controler("Erreur type du salaire moyen", erreur_type, ws["E10"].value)
    controler("Valeur critique du test sur la moyenne", critique, ws["E12"].value)

    resume = {
        "test": "Test sur une moyenne · salaire inférieur à 44 000 $",
        "h0": "Le salaire moyen est égal à 44 000 $",
        "h1": "Le salaire moyen est inférieur à 44 000 $",
        "seuil": seuil,
        "statistique": z,
        "valeur_critique": -invPhi(1 - seuil),
        "p_valeur": p_valeur,
        "decision": "rejet de H0" if moyenne < critique else "H0 non rejetée",
        "statistique_classeur": None,
        "decision_classeur": "rejet de H0",
    }

    portee = pd.DataFrame(
        [
            {
                "groupe": f"Manœuvres de {experience_min} années d'expérience et plus",
                "n": n,
                "salaire_moyen": moyenne,
                "ecart_type": ecart_type,
            },
            {
                "groupe": "Ensemble des manœuvres échantillonnés",
                "n": len(beaucerons),
                "salaire_moyen": beaucerons["salaire"].mean(),
                "ecart_type": beaucerons["salaire"].std(ddof=1),
            },
        ]
    )
    portee["corr_salaire_experience"] = beaucerons["salaire"].corr(beaucerons["experience"])
    return portee, resume


# =========================================================================
# 9. Test sur une proportion : pièces non conformes de Kansas Vamal
# =========================================================================
def test_proportion(pieces, wb):
    ws = wb["Gabarit"]
    seuil = ws["B18"].value
    proportion_reference = ws["B21"].value

    n = len(pieces)
    non_conformes = sum(1 for p in pieces if p == 2)
    p_observee = non_conformes / n
    erreur_type = math.sqrt(
        proportion_reference * (1 - proportion_reference) / n
    )
    z_seuil = invPhi(1 - seuil)
    critique = proportion_reference + z_seuil * erreur_type
    z = (p_observee - proportion_reference) / erreur_type
    p_valeur = 1 - Phi(z)

    controler("Pièces non conformes dénombrées", non_conformes, ws["B5"].value, 0)
    controler("Proportion observée de non-conformes", p_observee, ws["C5"].value)
    controler("Erreur type de la proportion", erreur_type * 100, ws["B31"].value)
    controler("Valeur critique du test sur la proportion", critique * 100, ws["B33"].value)

    resume = {
        "test": "Test sur une proportion · pièces non conformes",
        "h0": "La proportion de pièces non conformes est de 0,5 %",
        "h1": "La proportion de pièces non conformes dépasse 0,5 %",
        "seuil": seuil,
        "statistique": z,
        "valeur_critique": z_seuil,
        "p_valeur": p_valeur,
        "decision": "rejet de H0" if p_observee > critique else "H0 non rejetée",
        "statistique_classeur": None,
        "decision_classeur": "rejet de H0",
    }

    # Sensibilité : à partir de combien de pièces non conformes la décision
    # bascule-t-elle, et que devient-elle si le seuil change ?
    lignes = []
    for alpha in [0.01, 0.02, 0.05, 0.06, 0.10]:
        borne = proportion_reference + invPhi(1 - alpha) * erreur_type
        bascule = math.floor(borne * n) + 1
        lignes.append(
            {
                "seuil": alpha,
                "proportion_critique": borne,
                "proportion_observee": p_observee,
                "pieces_necessaires": bascule,
                "pieces_observees": non_conformes,
                "decision": "rejet de H0" if p_observee > borne else "H0 non rejetée",
            }
        )
    return pd.DataFrame(lignes), resume


# =========================================================================
# Programme principal
# =========================================================================
def main():
    # La console Windows utilise cp1252 par défaut : les libellés accentués et
    # les symboles mathématiques du rapport de contrôle y sont illisibles.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    DOSSIER_DONNEES.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(CLASSEUR, data_only=True)

    abus, beaucerons, pieces = lire_echantillons(wb)
    abus.to_csv(DOSSIER_DONNEES / "echantillon_abus.csv", index=False)
    beaucerons.to_csv(DOSSIER_DONNEES / "echantillon_beaucerons.csv", index=False)

    distribution, distribution_ronde, mesures = analyser_anciennete(abus, wb)
    distribution.to_csv(DOSSIER_DONNEES / "distribution_anciennete.csv", index=False)
    distribution_ronde.to_csv(
        DOSSIER_DONNEES / "distribution_anciennete_bornes_rondes.csv", index=False
    )
    mesures.to_csv(DOSSIER_DONNEES / "mesures_anciennete.csv", index=False)

    actes, profil = analyser_rebellion(abus, wb)
    actes.to_csv(DOSSIER_DONNEES / "actes_rebellion.csv", index=False)
    profil.to_csv(DOSSIER_DONNEES / "profil_rebellion.csv", index=False)

    regression = analyser_regression(abus, wb)
    regression.to_csv(DOSSIER_DONNEES / "regression_absences.csv", index=False)
    qualite_absences(abus, regression).to_csv(
        DOSSIER_DONNEES / "qualite_absences.csv", index=False
    )

    croise, detail_khi2, resume_independance = test_independance(abus, wb)
    croise.to_csv(DOSSIER_DONNEES / "tableau_croise_sexe.csv", index=False)
    detail_khi2.to_csv(DOSSIER_DONNEES / "detail_khi2_independance.csv", index=False)

    ic = intervalle_confiance(abus, wb)
    ic.to_csv(DOSSIER_DONNEES / "intervalle_confiance.csv", index=False)

    salaires, ecarts, resume_normalite = test_normalite(beaucerons, wb)
    salaires.to_csv(DOSSIER_DONNEES / "distribution_salaires.csv", index=False)
    ecarts.to_csv(DOSSIER_DONNEES / "parametres_salaires.csv", index=False)

    portee, resume_moyenne = test_moyenne(beaucerons, wb)
    portee.to_csv(DOSSIER_DONNEES / "portee_test_moyenne.csv", index=False)

    sensibilite, resume_proportion = test_proportion(pieces, wb)
    sensibilite.to_csv(DOSSIER_DONNEES / "sensibilite_proportion.csv", index=False)

    tests = pd.DataFrame(
        [resume_independance, resume_normalite, resume_moyenne, resume_proportion]
    )
    tests.to_csv(DOSSIER_DONNEES / "tests_hypotheses.csv", index=False)

    tableau_controles = pd.DataFrame(controles)
    tableau_controles.to_csv(DOSSIER_DONNEES / "controles.csv", index=False)

    largeur = max(len(c["controle"]) for c in controles)
    print("\nCONTRÔLES CROISÉS (recalcul indépendant vs valeur du classeur)\n")
    for c in controles:
        valeur = "  (aucune)" if c["classeur"] is None else f"{c['classeur']:>14,.6f}"
        marque = "  " if c["statut"] == "ok" else "!!"
        print(
            f"{marque} {c['controle']:<{largeur}}  "
            f"{c['recalcule']:>14,.6f}  {valeur}  {c['statut']}"
        )
    ecarts = sum(1 for c in controles if c["statut"] != "ok")
    print(f"\n{len(controles)} contrôles · {ecarts} écart(s)\n")

    print("Fichiers écrits dans", DOSSIER_DONNEES)


if __name__ == "__main__":
    main()
