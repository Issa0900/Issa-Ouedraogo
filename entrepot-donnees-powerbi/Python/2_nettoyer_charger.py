"""
2_nettoyer_charger.py — Nettoie les 8 sources brutes et charge un entrepot en etoile.

Principe directeur : une ligne douteuse n'est jamais supprimee en silence. Elle est
soit CORRIGEE (le defaut est reparable : format, casse, espace, variante d'ecriture),
soit REJETEE (le defaut rend la ligne ininterpretable : date illisible, article
inexistant, quantite negative sur une vente). Dans les deux cas la decision est
tracee dans la table `qualite_rejets`, avec la valeur d'origine.

C'est ce qui permet au script 3 de produire un rapport qualite verifiable et au
tableau de bord Power BI d'afficher un onglet "qualite des donnees" honnete.

Sortie : data/entrepot/boreal.db (SQLite, schema en etoile).
"""

from __future__ import annotations

import csv
import json
import os
import re
import sqlite3
import unicodedata
from datetime import date, datetime, timedelta

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BRUT = os.path.join(BASE, "data", "brut")
ENTREPOT = os.path.join(BASE, "data", "entrepot")
os.makedirs(ENTREPOT, exist_ok=True)
DB = os.path.join(ENTREPOT, "boreal.db")

# Encodage reel de chaque source : l'ERP exporte en cp1252, la comptabilite en
# latin-1, les systemes plus recents en UTF-8. Se tromper ici corrompt les accents
# sans lever d'erreur — c'est le premier piege du pipeline.
ENCODAGES = {
    "erp_ventes_2024.csv": "cp1252",
    "erp_ventes_2025.csv": "cp1252",
    "compta_charges.csv": "latin-1",
    "achats_fournisseurs.csv": "utf-8",
    "stock_inventaire.csv": "utf-8",
    "catalogue_produits.csv": "utf-8",
    "fournisseurs.csv": "utf-8",
    "entrepots.csv": "utf-8",
}

# Journal qualite : (source, cle, champ, valeur d'origine, motif, action)
rejets: list[tuple] = []


def tracer(source: str, cle: str, champ: str, valeur, motif: str, action: str) -> None:
    rejets.append((source, str(cle)[:60], champ, str(valeur)[:80], motif, action))


# ---------------------------------------------------------------------------
# Primitives de nettoyage
# ---------------------------------------------------------------------------

ESPACES = "   \t"  # insecable, insecable fine, fine, tabulation


def txt(valeur) -> str:
    """Supprime les espaces parasites (y compris insecables) et normalise les blancs."""
    if valeur is None:
        return ""
    s = str(valeur)
    for e in ESPACES:
        s = s.replace(e, " ")
    return re.sub(r"\s+", " ", s).strip()


def sans_accent(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def montant(valeur):
    """
    '1 234,56 $' / '1\\xa0234,56\\xa0$' / '1234.56' / 1234.56  ->  1234.56

    Les exports francophones melangent symbole monetaire, espace insecable comme
    separateur de milliers et virgule decimale. Aucun de ces trois elements n'est
    gere par float() : c'est la conversion la plus frequemment ratee du pipeline.
    """
    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, (int, float)):
        return float(valeur)
    s = txt(valeur).replace("$", "").replace(" ", "")
    if s.count(",") and s.count("."):
        # format anglais '1,234.56' : la virgule est un separateur de milliers
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


MOIS_FR = {"janv": 1, "fevr": 2, "mars": 3, "avr": 4, "mai": 5, "juin": 6,
           "juil": 7, "aout": 8, "sept": 9, "oct": 10, "nov": 11, "dec": 12}
FORMATS = ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y")


def parser_date(valeur):
    """
    Accepte les trois formats sortis par l'ERP ('05/03/2024', '2024-03-05',
    '05-mars-24') ainsi que les objets date/datetime deja types venant d'Excel.
    Retourne None si la valeur est inexploitable — jamais une date inventee.
    """
    if valeur is None or valeur == "":
        return None
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    s = txt(valeur)
    if not s or s.lower() in {"n/d", "n/a", "nd", "00/00/0000", "-"}:
        return None
    for f in FORMATS:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            pass
    # format '05-mars-24' : mois abrege en francais, annee sur deux chiffres
    m = re.match(r"^(\d{1,2})-([a-zA-ZéûôA-Z]+)\.?-(\d{2,4})$", s)
    if m:
        mois = MOIS_FR.get(sans_accent(m.group(2)).lower()[:4].rstrip("."))
        if mois is None:
            mois = MOIS_FR.get(sans_accent(m.group(2)).lower()[:3])
        if mois:
            annee = int(m.group(3))
            annee += 2000 if annee < 100 else 0
            try:
                return date(annee, mois, int(m.group(1)))
            except ValueError:
                return None
    return None


PROVINCES = {
    "qc": "QC", "quebec": "QC", "québec": "QC",
    "on": "ON", "ontario": "ON",
    "nb": "NB", "nouveau-brunswick": "NB",
    "ns": "NS", "nouvelle-ecosse": "NS", "nouvelle-écosse": "NS",
    "pe": "PE", "ile-du-prince-edouard": "PE", "île-du-prince-édouard": "PE",
}


def normaliser_province(valeur):
    """Ramene 'Québec' / 'Quebec' / 'qc' / 'QC' a un unique code ISO."""
    s = txt(valeur).lower()
    return PROVINCES.get(s) or PROVINCES.get(sans_accent(s))


# Particules qui restent en minuscules a l'interieur d'un toponyme :
# 'Saint-Jean-sur-Richelieu', 'Val-d\'Or', 'Baie-du-Febvre'.
PARTICULES = {"sur", "sous", "du", "de", "des", "la", "le", "les", "aux", "au", "et", "d"}


def normaliser_ville(valeur: str) -> str:
    """
    'MONTREAL' / 'montreal ' / 'Montréal ' -> 'Montréal'

    Decoupe sur l'espace, le trait d'union ET l'apostrophe : un simple .title()
    produirait 'Thunder bay' (un seul mot capitalise), 'Saint-Jean-Sur-Richelieu'
    (particule capitalisee) ou "Val-D'or" (apostrophe ignoree).
    """
    s = txt(valeur)
    if not s:
        return ""
    morceaux = re.split(r"([ \-'])", s)
    sortie = []
    for i, bloc in enumerate(morceaux):
        if bloc in (" ", "-", "'"):
            sortie.append(bloc)
        elif i > 0 and bloc.lower() in PARTICULES:
            sortie.append(bloc.lower())
        else:
            sortie.append(bloc[:1].upper() + bloc[1:].lower() if bloc else bloc)
    return "".join(sortie)


COURRIEL = re.compile(r"^[^@\s]+@[^@\s]+\.[a-z]{2,}$", re.I)


def normaliser_courriel(valeur):
    s = txt(valeur).lower()
    return s if COURRIEL.match(s) else None


def cle_entreprise(nom: str) -> str:
    """Cle de rapprochement : sans accent, sans ponctuation, sans forme juridique."""
    s = sans_accent(txt(nom)).upper()
    s = re.sub(r"[^A-Z0-9 ]", "", s)
    s = re.sub(r"\b(INC|LTEE|LTD|ENR|ET FILS)\b", "", s)
    return re.sub(r"\s+", " ", s).strip()


def lire_csv(fichier: str, delimiteur: str = ","):
    chemin = os.path.join(BRUT, fichier)
    with open(chemin, encoding=ENCODAGES.get(fichier, "utf-8"), newline="") as f:
        yield from csv.DictReader(f, delimiter=delimiteur)


def est_ligne_technique(ligne: dict) -> bool:
    """Pied de page ajoute par l'outil d'export ('TOTAL GENERAL', 'Export du ...')."""
    premiere = txt(next(iter(ligne.values()), ""))
    if not premiere:
        return all(not txt(v) for v in ligne.values())
    return bool(re.match(r"^(TOTAL|Export du|Sous-total)", premiere, re.I))


# ---------------------------------------------------------------------------
# Dimensions
# ---------------------------------------------------------------------------

# --- dim_entrepot -----------------------------------------------------------
entrepots = {}
for i, r in enumerate(lire_csv("entrepots.csv"), start=1):
    entrepots[txt(r["code"])] = {
        "id": i, "code": txt(r["code"]), "nom": txt(r["nom"]),
        "region": txt(r["region"]), "surface_m2": int(r["surface_m2"]),
    }

# --- dim_fournisseur --------------------------------------------------------
fournisseurs = {}
for i, r in enumerate(lire_csv("fournisseurs.csv"), start=1):
    fournisseurs[txt(r["code"])] = {
        "id": i, "code": txt(r["code"]), "nom": txt(r["nom"]), "pays": txt(r["pays"]),
        "delai_moyen_jours": int(r["delai_moyen_jours"]), "conditions": txt(r["conditions"]),
    }

# --- dim_produit ------------------------------------------------------------
produits = {}
for r in lire_csv("catalogue_produits.csv"):
    sku = txt(r["sku"])
    if sku in produits:
        # Le meme SKU apparait deux fois au catalogue : on garde la premiere fiche.
        tracer("catalogue_produits.csv", sku, "sku", sku, "doublon_exact", "rejetee")
        continue
    cout = montant(r["cout_standard"])
    if not cout:
        # Sans cout standard, la marge theorique du produit est incalculable. La fiche
        # reste utilisable (les ventes portent leur propre cout reel) mais on le signale.
        tracer("catalogue_produits.csv", sku, "cout_standard", r["cout_standard"],
               "cout_standard_absent", "signalee")
        cout = None
    produits[sku] = {
        "id": len(produits) + 1, "sku": sku, "designation": txt(r["designation"]),
        "famille": txt(r["famille"]), "fournisseur_code": txt(r["code_fournisseur"]),
        "cout_standard": cout, "prix_liste": montant(r["prix_liste"]),
        "poids_kg": montant(r["poids_kg"]),
    }

# --- dim_client (JSON imbrique + rapprochement des doublons metier) ---------
with open(os.path.join(BRUT, "crm_clients.json"), encoding="utf-8") as f:
    fiches_crm = json.load(f)["clients"]

# Les corrections de champ sont d'abord mises de cote, puis journalisees uniquement
# pour la fiche finalement conservee : consigner une correction sur une fiche qui
# sera fusionnee reviendrait a compter deux fois le meme defaut.
fiches = []
for fiche in fiches_crm:
    ident, adr, com = fiche["identification"], fiche["adresse"], fiche["commercial"]
    code = txt(ident["code"])
    en_attente = []

    province = normaliser_province(adr["province"])
    if province is None:
        en_attente.append(("province", adr["province"], "province_inconnue", "signalee"))
    elif txt(adr["province"]) != province:
        en_attente.append(("province", adr["province"], "variante_ecriture", "corrigee"))

    ville_brute = adr["ville"]
    ville = normaliser_ville(ville_brute)
    if ville_brute != ville:
        en_attente.append(("ville", ville_brute, "casse_ou_espaces", "corrigee"))

    courriel = normaliser_courriel(com["courriel"])
    if courriel is None:
        motif = "courriel_invalide" if txt(com["courriel"]) else "valeur_manquante"
        en_attente.append(("courriel", com["courriel"], motif, "corrigee"))

    cp = txt(adr["code_postal"]).upper() or None
    if cp is None:
        en_attente.append(("code_postal", adr["code_postal"], "valeur_manquante", "corrigee"))

    fiches.append({
        "_en_attente": en_attente,
        "code": code, "nom": txt(ident["raison_sociale"]), "type": txt(ident["type_commerce"]),
        "date_ouverture": parser_date(ident["date_ouverture_compte"]),
        "ville": ville, "province": province, "code_postal": cp,
        "segment": txt(com["segment"]), "conditions_paiement": txt(com["conditions_paiement"]),
        "representant": txt(com["representant_matricule"]), "courriel": courriel,
        "cle": cle_entreprise(ident["raison_sociale"]),
    })

# Rapprochement : deux fiches decrivent la meme entreprise si la raison sociale
# normalisee ET le courriel concordent (ou si aucune des deux n'a de courriel).
# On conserve la fiche au code le plus ancien et on redirige les ventes des autres.
groupes: dict[tuple, list] = {}
for f_ in fiches:
    groupes.setdefault((f_["cle"], f_["courriel"]), []).append(f_)

clients = {}
alias_client = {}   # code secondaire -> code conserve
for (cle, _), membres in groupes.items():
    membres.sort(key=lambda x: x["code"])
    principal = membres[0]
    principal["id"] = len(clients) + 1
    principal["fiches_fusionnees"] = len(membres)
    clients[principal["code"]] = principal
    for champ, valeur, motif, action in principal["_en_attente"]:
        tracer("crm_clients.json", principal["code"], champ, valeur, motif, action)
    alias_client[principal["code"]] = principal["code"]
    for doublon in membres[1:]:
        alias_client[doublon["code"]] = principal["code"]
        tracer("crm_clients.json", doublon["code"], "raison_sociale", doublon["nom"],
               "doublon_metier", f"fusionnee vers {principal['code']}")

# --- dim_employe (Excel : lignes de titre avant l'en-tete, salaires en texte) ---
from openpyxl import load_workbook

wb = load_workbook(os.path.join(BRUT, "rh_employes.xlsx"), data_only=True)
ws = wb["Employés"]
grille = list(ws.iter_rows(values_only=True))

# L'en-tete n'est pas en ligne 1 : on le localise plutot que de le supposer.
ligne_entete = next(i for i, r in enumerate(grille) if r and txt(r[0]) == "Matricule")
colonnes = [txt(c) for c in grille[ligne_entete]]
IDX = {nom: i for i, nom in enumerate(colonnes) if nom}

employes = {}
for r in grille[ligne_entete + 1:]:
    if not r or not txt(r[IDX["Matricule"]]):
        continue
    matricule = txt(r[IDX["Matricule"]])
    if matricule.upper().startswith("TOTAL"):
        tracer("rh_employes.xlsx", matricule, "-", "ligne de totaux", "ligne_technique", "rejetee")
        continue
    if matricule in employes:
        tracer("rh_employes.xlsx", matricule, "matricule", matricule, "doublon_exact", "rejetee")
        continue
    salaire_brut = r[IDX["Salaire annuel"]]
    salaire = montant(salaire_brut)
    if isinstance(salaire_brut, str):
        tracer("rh_employes.xlsx", matricule, "Salaire annuel", salaire_brut,
               "nombre_en_texte", "corrigee")
    depart = parser_date(r[IDX["Date de départ"]])
    employes[matricule] = {
        "id": len(employes) + 1, "matricule": matricule,
        "prenom": txt(r[IDX["Prénom"]]), "nom": txt(r[IDX["Nom"]]),
        "nom_complet": f"{txt(r[IDX['Prénom']])} {txt(r[IDX['Nom']])}",
        "poste": txt(r[IDX["Poste"]]), "departement": txt(r[IDX["Département"]]),
        "entrepot_code": txt(r[IDX["Entrepôt"]]),
        "date_embauche": parser_date(r[IDX["Date d'embauche"]]),
        "date_depart": depart, "salaire_annuel": salaire,
        "statut": txt(r[IDX["Statut"]]), "est_actif": 0 if depart else 1,
    }

# --- dim_date ---------------------------------------------------------------
JOURS = ["lundi", "mardi", "mercredi", "jeudi", "vendredi", "samedi", "dimanche"]
MOIS_NOM = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet",
            "août", "septembre", "octobre", "novembre", "décembre"]

DEBUT_CAL, FIN_CAL = date(2024, 1, 1), date(2026, 6, 30)
calendrier = []
j = DEBUT_CAL
while j <= FIN_CAL:
    calendrier.append({
        "date_id": int(j.strftime("%Y%m%d")), "date": j.isoformat(), "annee": j.year,
        "trimestre": f"T{(j.month - 1)//3 + 1}", "mois": j.month,
        "nom_mois": MOIS_NOM[j.month - 1], "mois_annee": j.strftime("%Y-%m"),
        "semaine": int(j.strftime("%V")), "jour_semaine": j.weekday() + 1,
        "nom_jour": JOURS[j.weekday()], "est_weekend": 1 if j.weekday() >= 5 else 0,
    })
    j += timedelta(days=1)

did = lambda d: int(d.strftime("%Y%m%d")) if d else None


# ---------------------------------------------------------------------------
# Tables de faits
# ---------------------------------------------------------------------------

# --- fait_ventes ------------------------------------------------------------
# Seuil de detection des erreurs de saisie : un prix de vente ne peut pas depasser
# le prix de liste (il n'y a que des remises, jamais de majoration). On tolere un
# facteur 3 avant de conclure a une erreur, pour ne pas rejeter un cas legitime.
FACTEUR_PRIX_ABERRANT = 3.0

ventes = []
vues = set()
for fichier in ("erp_ventes_2024.csv", "erp_ventes_2025.csv"):
    for r in lire_csv(fichier, delimiteur=";"):
        if est_ligne_technique(r):
            tracer(fichier, txt(r.get("NO COMMANDE")), "-", "pied de page d'export",
                   "ligne_technique", "rejetee")
            continue

        empreinte = tuple(txt(v) for v in r.values())
        if empreinte in vues:
            tracer(fichier, txt(r["NO COMMANDE"]), "-", "ligne identique",
                   "doublon_exact", "rejetee")
            continue
        vues.add(empreinte)

        cle = txt(r["NO COMMANDE"])
        d_cmd = parser_date(r["DATE COMMANDE"])
        if d_cmd is None:
            tracer(fichier, cle, "DATE COMMANDE", r["DATE COMMANDE"], "date_illisible", "rejetee")
            continue

        sku = txt(r["CODE ARTICLE"]).upper()
        produit = produits.get(sku)
        if produit is None:
            tracer(fichier, cle, "CODE ARTICLE", r["CODE ARTICLE"], "article_inconnu", "rejetee")
            continue

        code_brut = r["CODE CLIENT"]
        code_normalise = txt(code_brut).upper()
        code_client = alias_client.get(code_normalise)
        if code_client is None:
            tracer(fichier, cle, "CODE CLIENT", code_brut, "client_inconnu", "rejetee")
            continue
        if code_normalise != code_client:
            # La vente pointait vers une fiche CRM doublon : elle est rattachee au
            # client conserve, sinon le chiffre d'affaires serait eclate en deux.
            tracer(fichier, cle, "CODE CLIENT", code_brut, "client_fusionne",
                   f"rattachee a {code_client}")
        elif code_brut != code_normalise:
            tracer(fichier, cle, "CODE CLIENT", code_brut, "casse_ou_espaces", "corrigee")

        try:
            qte = int(float(txt(r["QTE"])))
        except (ValueError, TypeError):
            qte = 0
        if qte <= 0:
            tracer(fichier, cle, "QTE", r["QTE"], "quantite_invalide", "rejetee")
            continue

        prix = montant(r["PRIX UNITAIRE"])
        if not prix or prix <= 0:
            tracer(fichier, cle, "PRIX UNITAIRE", r["PRIX UNITAIRE"], "prix_invalide", "rejetee")
            continue
        if produit["prix_liste"] and prix > produit["prix_liste"] * FACTEUR_PRIX_ABERRANT:
            tracer(fichier, cle, "PRIX UNITAIRE", r["PRIX UNITAIRE"], "prix_aberrant", "rejetee")
            continue

        canal = txt(r["CANAL DE VENTE"])
        if not canal or canal.upper() in {"N/A", "ND"}:
            tracer(fichier, cle, "CANAL DE VENTE", r["CANAL DE VENTE"],
                   "valeur_manquante", "corrigee")
            canal = "Non renseigné"

        matricule = txt(r["REPRESENTANT"])
        employe = employes.get(matricule)
        if employe is None and matricule:
            tracer(fichier, cle, "REPRESENTANT", matricule, "employe_inconnu", "corrigee")

        d_liv = parser_date(r["DATE LIVRAISON"])
        d_pai = parser_date(r["DATE PAIEMENT"])
        cout = montant(r["COUT UNITAIRE"]) or 0.0
        montant_ht = round(qte * prix, 2)
        cout_total = round(qte * cout, 2)

        ventes.append({
            "no_commande": cle,
            "date_commande_id": did(d_cmd), "date_livraison_id": did(d_liv),
            "date_paiement_id": did(d_pai),
            "client_id": clients[code_client]["id"], "produit_id": produit["id"],
            "entrepot_id": entrepots[txt(r["ENTREPOT"])]["id"],
            "employe_id": employe["id"] if employe else None,
            "quantite": qte, "prix_unitaire": prix,
            "remise_pct": montant(r["REMISE %"]), "cout_unitaire": cout,
            "montant_ht": montant_ht, "cout_total": cout_total,
            "marge": round(montant_ht - cout_total, 2),
            "canal_vente": canal,
            "delai_paiement_jours": (d_pai - d_cmd).days if d_pai else None,
            "delai_livraison_jours": (d_liv - d_cmd).days if d_liv else None,
        })

# --- fait_achats ------------------------------------------------------------
achats = []
for r in lire_csv("achats_fournisseurs.csv", delimiteur="\t"):
    cle = txt(r["no_achat"])
    sku = txt(r["code_article"]).upper()
    produit = produits.get(sku)
    if produit is None:
        tracer("achats_fournisseurs.csv", cle, "code_article", r["code_article"],
               "article_inconnu", "rejetee")
        continue
    d_cmd = parser_date(r["date_commande"])
    if d_cmd is None:
        tracer("achats_fournisseurs.csv", cle, "date_commande", r["date_commande"],
               "date_illisible", "rejetee")
        continue
    qte = int(float(txt(r["quantite"])))
    # Une quantite negative sur un achat n'est pas une erreur : c'est un retour au
    # fournisseur. On la conserve en la qualifiant, au lieu de la rejeter.
    est_retour = 1 if qte < 0 else 0
    if est_retour:
        tracer("achats_fournisseurs.csv", cle, "quantite", r["quantite"],
               "quantite_negative", "conservee comme retour")
    cout = montant(r["cout_unitaire"]) or 0.0
    d_prevue, d_recue = parser_date(r["date_prevue"]), parser_date(r["date_reception"])
    achats.append({
        "no_achat": cle, "date_commande_id": did(d_cmd), "date_prevue_id": did(d_prevue),
        "date_reception_id": did(d_recue),
        "fournisseur_id": fournisseurs[txt(r["code_fournisseur"])]["id"],
        "produit_id": produit["id"], "quantite": qte, "cout_unitaire": cout,
        "montant": round(qte * cout, 2), "statut": txt(r["statut"]),
        "est_retour": est_retour,
        "retard_jours": (d_recue - d_prevue).days if (d_recue and d_prevue) else None,
    })

# --- fait_stock -------------------------------------------------------------
stocks = []
photos_vues = set()
for r in lire_csv("stock_inventaire.csv"):
    sku = txt(r["code_article"]).upper()
    cle = f"{r['date_photo']}|{r['code_entrepot']}|{sku}"
    produit = produits.get(sku)
    if produit is None:
        tracer("stock_inventaire.csv", cle, "code_article", r["code_article"],
               "article_inconnu", "rejetee")
        continue
    if cle in photos_vues:
        tracer("stock_inventaire.csv", cle, "-", "photo deja enregistree",
               "doublon_exact", "rejetee")
        continue
    photos_vues.add(cle)
    qte = int(float(txt(r["quantite_en_stock"])))
    if qte < 0:
        # Un comptage d'inventaire ne peut pas etre negatif : la ligne est inexploitable.
        tracer("stock_inventaire.csv", cle, "quantite_en_stock", r["quantite_en_stock"],
               "quantite_invalide", "rejetee")
        continue
    d = parser_date(r["date_photo"])
    cout = montant(r["cout_unitaire"]) or 0.0
    seuil = int(float(txt(r["seuil_minimum"])))
    stocks.append({
        "date_id": did(d), "entrepot_id": entrepots[txt(r["code_entrepot"])]["id"],
        "produit_id": produit["id"], "quantite": qte, "cout_unitaire": cout,
        "valeur_stock": round(qte * cout, 2), "seuil_minimum": seuil,
        "sous_seuil": 1 if qte < seuil else 0,
    })

# --- fait_paie --------------------------------------------------------------
paie = []
for r in wb["Paie mensuelle"].iter_rows(min_row=2, values_only=True):
    if not r or not r[1]:
        continue
    matricule = txt(r[1])
    employe = employes.get(matricule)
    if employe is None:
        tracer("rh_employes.xlsx", matricule, "Matricule", matricule,
               "employe_inconnu", "rejetee")
        continue
    d = parser_date(txt(r[0]) + "-01")
    paie.append({
        "date_id": did(d), "employe_id": employe["id"],
        "salaire_base": montant(r[2]), "primes": montant(r[3]),
        "heures_supplementaires": montant(r[4]), "charges_sociales": montant(r[5]),
        "cout_total": montant(r[6]),
    })

# --- fait_marketing ---------------------------------------------------------
with open(os.path.join(BRUT, "marketing_campagnes.json"), encoding="utf-8") as f:
    campagnes = json.load(f)["campagnes"]

canaux = {}
marketing = []
for c in campagnes:
    canal = txt(c["canal"])
    if canal not in canaux:
        canaux[canal] = {"id": len(canaux) + 1, "canal": canal}
    d = parser_date(c["periode"][:10])
    budget = montant(c["budget_prevu"])
    if budget is None:
        tracer("marketing_campagnes.json", f"{c['periode'][:7]} {canal}", "budget_prevu",
               c["budget_prevu"], "valeur_manquante", "corrigee")
    depense = montant(c["depense_reelle"]) or 0.0
    marketing.append({
        "date_id": did(d), "canal_id": canaux[canal]["id"], "budget": budget,
        "depense": depense, "prospects": int(c["prospects_generes"]),
        "nouveaux_clients": int(c["nouveaux_clients"]),
        "ecart_budget": round(depense - budget, 2) if budget is not None else None,
    })

# --- fait_charges -----------------------------------------------------------
# La comptabilite saisit la meme categorie tantot en majuscules, tantot en
# minuscules. On ramene chaque ecriture a un libelle unique via une table de
# correspondance, plutot que par une regle de casse qui casserait "Logiciels et TI".
CATEGORIES_CHARGES = ["Loyer", "Énergie", "Assurances", "Entretien et réparations",
                      "Télécommunications", "Logiciels et TI", "Honoraires professionnels",
                      "Fournitures de bureau", "Transport et livraison"]
REF_CHARGES = {sans_accent(c).lower(): c for c in CATEGORIES_CHARGES}

charges = []
ecritures_vues = set()
for r in lire_csv("compta_charges.csv", delimiteur=";"):
    brute = txt(r["Categorie"])
    categorie = REF_CHARGES.get(sans_accent(brute).lower())
    if categorie is None:
        tracer("compta_charges.csv", f"{r['Mois']} {brute}", "Categorie", brute,
               "categorie_inconnue", "rejetee")
        continue
    mois = txt(r["Mois"])
    cle = f"{mois}|{categorie}|{txt(r['Entrepot'])}"
    if cle in ecritures_vues:
        tracer("compta_charges.csv", cle, "-", "ecriture deja comptabilisee",
               "doublon_exact", "rejetee")
        continue
    ecritures_vues.add(cle)
    # Journalise apres le dedoublonnage : une ecriture rejetee n'a pas a figurer
    # aussi dans le decompte des libelles normalises.
    if brute != categorie:
        tracer("compta_charges.csv", f"{mois} {brute}", "Categorie", brute,
               "variante_ecriture", "corrigee")
    d = parser_date(f"01/{mois}")
    charges.append({
        "date_id": did(d), "categorie": categorie,
        "entrepot_id": entrepots[txt(r["Entrepot"])]["id"],
        "montant": montant(r["Montant"]),
    })


# ---------------------------------------------------------------------------
# Chargement dans l'entrepot SQLite
# ---------------------------------------------------------------------------
# Le DDL vit dans SQL/1_creation_schema.sql plutot que dans des chaines Python :
# le schema doit rester lisible et rejouable sans passer par ce script.

with open(os.path.join(BASE, "SQL", "1_creation_schema.sql"), encoding="utf-8") as f:
    ddl = f.read()

if os.path.exists(DB):
    os.remove(DB)
cx = sqlite3.connect(DB)
cx.executescript(ddl)


def inserer(table: str, colonnes: list[str], lignes: list[dict]) -> int:
    if not lignes:
        return 0
    marqueurs = ",".join("?" * len(colonnes))
    cx.executemany(
        f"INSERT INTO {table} ({','.join(colonnes)}) VALUES ({marqueurs})",
        [[lg.get(c) for c in colonnes] for lg in lignes],
    )
    return len(lignes)


COLS_DATE = ["date_id", "date", "annee", "trimestre", "mois", "nom_mois", "mois_annee",
             "semaine", "jour_semaine", "nom_jour", "est_weekend"]
inserer("dim_date", COLS_DATE, calendrier)

inserer("dim_entrepot", ["entrepot_id", "code", "nom", "region", "surface_m2"],
        [{**e, "entrepot_id": e["id"]} for e in entrepots.values()])

inserer("dim_fournisseur",
        ["fournisseur_id", "code", "nom", "pays", "delai_moyen_jours", "conditions"],
        [{**f_, "fournisseur_id": f_["id"]} for f_ in fournisseurs.values()])

inserer("dim_produit",
        ["produit_id", "sku", "designation", "famille", "fournisseur_code",
         "cout_standard", "prix_liste", "poids_kg"],
        [{**p, "produit_id": p["id"]} for p in produits.values()])

inserer("dim_client",
        ["client_id", "code", "nom", "type_commerce", "ville", "province", "code_postal",
         "segment", "conditions_paiement", "courriel", "date_ouverture", "fiches_fusionnees"],
        [{**c, "client_id": c["id"], "type_commerce": c["type"],
          "date_ouverture": c["date_ouverture"].isoformat() if c["date_ouverture"] else None}
         for c in clients.values()])

inserer("dim_employe",
        ["employe_id", "matricule", "prenom", "nom", "nom_complet", "poste", "departement",
         "entrepot_code", "date_embauche", "date_depart", "salaire_annuel", "statut", "est_actif"],
        [{**e, "employe_id": e["id"],
          "date_embauche": e["date_embauche"].isoformat() if e["date_embauche"] else None,
          "date_depart": e["date_depart"].isoformat() if e["date_depart"] else None}
         for e in employes.values()])

inserer("dim_canal_marketing", ["canal_id", "canal"],
        [{"canal_id": c["id"], "canal": c["canal"]} for c in canaux.values()])

inserer("fait_ventes",
        ["no_commande", "date_commande_id", "date_livraison_id", "date_paiement_id",
         "client_id", "produit_id", "entrepot_id", "employe_id", "quantite", "prix_unitaire",
         "remise_pct", "cout_unitaire", "montant_ht", "cout_total", "marge", "canal_vente",
         "delai_paiement_jours", "delai_livraison_jours"], ventes)

inserer("fait_achats",
        ["no_achat", "date_commande_id", "date_prevue_id", "date_reception_id",
         "fournisseur_id", "produit_id", "quantite", "cout_unitaire", "montant", "statut",
         "est_retour", "retard_jours"], achats)

inserer("fait_stock",
        ["date_id", "entrepot_id", "produit_id", "quantite", "cout_unitaire",
         "valeur_stock", "seuil_minimum", "sous_seuil"], stocks)

inserer("fait_paie",
        ["date_id", "employe_id", "salaire_base", "primes", "heures_supplementaires",
         "charges_sociales", "cout_total"], paie)

inserer("fait_marketing",
        ["date_id", "canal_id", "budget", "depense", "prospects", "nouveaux_clients",
         "ecart_budget"], marketing)

inserer("fait_charges", ["date_id", "categorie", "entrepot_id", "montant"], charges)

cx.executemany(
    "INSERT INTO qualite_rejets (source, cle, champ, valeur, motif, action) "
    "VALUES (?,?,?,?,?,?)", rejets)

cx.commit()

# --- Controles d'integrite post-chargement ----------------------------------
# Un chargement qui « passe » n'est pas un chargement correct : on verifie que
# les jointures que Power BI va faire ne trouveront pas de cle orpheline.
controles = {
    "ventes sans date au calendrier":
        "SELECT COUNT(*) FROM fait_ventes v LEFT JOIN dim_date d "
        "ON v.date_commande_id = d.date_id WHERE d.date_id IS NULL",
    "ventes sans client":
        "SELECT COUNT(*) FROM fait_ventes v LEFT JOIN dim_client c "
        "ON v.client_id = c.client_id WHERE c.client_id IS NULL",
    "ventes sans produit":
        "SELECT COUNT(*) FROM fait_ventes v LEFT JOIN dim_produit p "
        "ON v.produit_id = p.produit_id WHERE p.produit_id IS NULL",
    "achats sans fournisseur":
        "SELECT COUNT(*) FROM fait_achats a LEFT JOIN dim_fournisseur f "
        "ON a.fournisseur_id = f.fournisseur_id WHERE f.fournisseur_id IS NULL",
    "stock sans produit":
        "SELECT COUNT(*) FROM fait_stock s LEFT JOIN dim_produit p "
        "ON s.produit_id = p.produit_id WHERE p.produit_id IS NULL",
    "paie sans employe":
        "SELECT COUNT(*) FROM fait_paie f LEFT JOIN dim_employe e "
        "ON f.employe_id = e.employe_id WHERE e.employe_id IS NULL",
    "marge incoherente (montant - cout)":
        "SELECT COUNT(*) FROM fait_ventes "
        "WHERE ABS(marge - (montant_ht - cout_total)) > 0.01",
}
echecs = 0
for libelle, requete in controles.items():
    n = cx.execute(requete).fetchone()[0]
    if n:
        echecs += 1
        print(f"  ECHEC  {libelle}: {n}")
if echecs == 0:
    print("Controles d'integrite : 7/7 passes (aucune cle orpheline)")

n_rejetees = sum(1 for r in rejets if r[5] == "rejetee")
n_corrigees = sum(1 for r in rejets if r[5] != "rejetee")
print(f"\nEntrepot ecrit : {os.path.relpath(DB, BASE)}")
print(f"  dim_date {len(calendrier):>6} | dim_client {len(clients):>4} | "
      f"dim_produit {len(produits):>4} | dim_employe {len(employes):>3}")
print(f"  fait_ventes {len(ventes):>6} | fait_achats {len(achats):>5} | "
      f"fait_stock {len(stocks):>6}")
print(f"  fait_paie {len(paie):>8} | fait_marketing {len(marketing):>3} | "
      f"fait_charges {len(charges):>4}")
print(f"  qualite_rejets {len(rejets):>5} ({n_rejetees} lignes rejetees, "
      f"{n_corrigees} corrigees ou signalees)")
cx.close()
