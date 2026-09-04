"""
3_controle_qualite.py — Rapport qualite verifiable.

Un pipeline qui affirme "les donnees ont ete nettoyees" ne prouve rien. Ce script
confronte deux sources independantes :

  - ce que le generateur a REELLEMENT injecte  (data/brut/_anomalies_injectees.json)
  - ce que l'ETL a REELLEMENT detecte          (table qualite_rejets de l'entrepot)

Tout ecart est affiche et doit etre explique, pas masque. Un ecart non explique
signifie qu'une anomalie est passee au travers du nettoyage.

Sortie : data/RAPPORT-QUALITE.md
"""

from __future__ import annotations

import csv
import json
import os
import sqlite3

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BRUT = os.path.join(BASE, "data", "brut")
DB = os.path.join(BASE, "data", "entrepot", "boreal.db")
SORTIE = os.path.join(BASE, "data", "RAPPORT-QUALITE.md")

milliers = lambda v: f"{v:,}".replace(",", " ")
pourcent = lambda v: f"{v:.1f}".replace(".", ",") + " %"

cx = sqlite3.connect(DB)
cx.row_factory = sqlite3.Row

injecte = json.load(open(os.path.join(BRUT, "_anomalies_injectees.json"), encoding="utf-8"))["anomalies"]
detecte = {m: n for m, n in cx.execute("SELECT motif, COUNT(*) FROM qualite_rejets GROUP BY motif")}

# Correspondance entre les anomalies injectees et les motifs du journal de l'ETL.
# Plusieurs anomalies de sources differentes convergent vers un meme motif : c'est
# voulu, le motif decrit la NATURE du defaut, pas le fichier d'ou il vient.
CORRESPONDANCE = [
    ("date_illisible", ["ventes_date_invalide"],
     "Dates inexploitables ('31/02/2024', '00/00/0000', vide) : la ligne est ecartee, "
     "aucune date n'est devinee."),
    ("article_inconnu", ["ventes_sku_orphelin", "achats_sku_orphelin", "stock_sku_orphelin"],
     "Reference article absente du catalogue : la ligne ne peut etre rattachee ni a "
     "une famille ni a un fournisseur."),
    ("quantite_invalide", ["ventes_quantite_invalide", "stock_quantite_negative"],
     "Quantite nulle ou negative sur une vente, ou comptage d'inventaire negatif."),
    ("prix_invalide", ["ventes_prix_zero"], "Prix de vente a zero."),
    ("prix_aberrant", ["ventes_prix_aberrant"],
     "Prix superieur a 3x le prix de liste : virgule oubliee a la saisie."),
    ("doublon_exact", ["ventes_doublons", "stock_doublons", "charges_doublons",
                       "catalogue_doublons", "rh_doublons"],
     "Ligne strictement identique a une ligne deja chargee (double import, script "
     "d'extraction relance)."),
    ("doublon_metier", ["clients_doublons_metier"],
     "Meme entreprise saisie sous deux codes clients : les fiches sont fusionnees et "
     "les ventes rattachees au code conserve."),
    ("ligne_technique", ["ventes_lignes_pied_de_page", "rh_ligne_total"],
     "Ligne de total ou pied de page ajoute par l'outil d'export."),
    ("valeur_manquante", ["ventes_canal_manquant", "clients_code_postal_manquant",
                          "marketing_budget_manquant", "clients_courriel_vide"],
     "Champ non renseigne : conserve a NULL ou libelle 'Non renseigne', jamais impute."),
    ("casse_ou_espaces", ["ventes_code_client_sale", "clients_ville_casse"],
     "Espaces parasites ou casse incoherente sur une cle ou un libelle."),
    ("variante_ecriture", ["clients_province_variantes", "charges_categorie_casse"],
     "Meme valeur ecrite de plusieurs facons ('QC' / 'Quebec' / 'Québec') : ramenee "
     "a une forme unique."),
    ("courriel_invalide", ["clients_courriel_invalide"],
     "Courriel non conforme ('achats@', 'n/a') : mis a NULL plutot que conserve tel quel."),
    ("nombre_en_texte", ["rh_salaire_en_texte"],
     "Montant saisi en texte ('52 000 $') dans le classeur RH."),
    ("cout_standard_absent", ["catalogue_cout_zero"],
     "Cout standard a zero au catalogue : la fiche est conservee, le defaut est signale."),
    ("quantite_negative", ["achats_quantite_negative"],
     "Quantite negative sur un achat : ce n'est pas une erreur mais un retour "
     "fournisseur, conserve et qualifie."),
]

# Ecarts connus et expliques. Un ecart absent de ce dictionnaire est signale comme
# anomalie du pipeline lui-meme.
ECARTS_ATTENDUS = {
    "nombre_en_texte": "Un salaire en texte figurait sur la ligne d'un employe en "
                       "double, ecartee plus tot comme doublon exact : le defaut de "
                       "format n'a donc jamais eu a etre corrige.",
}

lignes_reconciliation = []
ecarts_inexpliques = []
for motif, cles, explication in CORRESPONDANCE:
    attendu = sum(injecte.get(k, 0) for k in cles)
    trouve = detecte.get(motif, 0)
    ecart = trouve - attendu
    lignes_reconciliation.append((motif, attendu, trouve, ecart, explication))
    if ecart != 0 and motif not in ECARTS_ATTENDUS:
        ecarts_inexpliques.append((motif, attendu, trouve))

couverts = {k for _, cles, _ in CORRESPONDANCE for k in cles}
non_couverts = sorted(set(injecte) - couverts)

# --- Volumetrie par source --------------------------------------------------
def compter_lignes(fichier: str, encodage: str, delim: str) -> int:
    with open(os.path.join(BRUT, fichier), encoding=encodage, newline="") as f:
        return sum(1 for _ in csv.DictReader(f, delimiter=delim))

VOLUMETRIE = [
    ("ERP ventes (2 fichiers)", "CSV ';' cp1252",
     compter_lignes("erp_ventes_2024.csv", "cp1252", ";")
     + compter_lignes("erp_ventes_2025.csv", "cp1252", ";"),
     cx.execute("SELECT COUNT(*) FROM fait_ventes").fetchone()[0]),
    ("Achats fournisseurs", "CSV tabulation UTF-8",
     compter_lignes("achats_fournisseurs.csv", "utf-8", "\t"),
     cx.execute("SELECT COUNT(*) FROM fait_achats").fetchone()[0]),
    ("Inventaire", "CSV ',' UTF-8",
     compter_lignes("stock_inventaire.csv", "utf-8", ","),
     cx.execute("SELECT COUNT(*) FROM fait_stock").fetchone()[0]),
    ("Charges comptables", "CSV ';' latin-1",
     compter_lignes("compta_charges.csv", "latin-1", ";"),
     cx.execute("SELECT COUNT(*) FROM fait_charges").fetchone()[0]),
    ("Catalogue produits", "CSV ',' UTF-8",
     compter_lignes("catalogue_produits.csv", "utf-8", ","),
     cx.execute("SELECT COUNT(*) FROM dim_produit").fetchone()[0]),
    ("CRM clients", "JSON imbrique UTF-8",
     len(json.load(open(os.path.join(BRUT, "crm_clients.json"), encoding="utf-8"))["clients"]),
     cx.execute("SELECT COUNT(*) FROM dim_client").fetchone()[0]),
    ("Marketing", "JSON UTF-8",
     len(json.load(open(os.path.join(BRUT, "marketing_campagnes.json"), encoding="utf-8"))["campagnes"]),
     cx.execute("SELECT COUNT(*) FROM fait_marketing").fetchone()[0]),
]

from openpyxl import load_workbook
wb = load_workbook(os.path.join(BRUT, "rh_employes.xlsx"), data_only=True)
n_rh = sum(1 for r in wb["Employés"].iter_rows(min_row=5, values_only=True) if r and r[0])
VOLUMETRIE.append(("RH employes", "XLSX 2 feuilles", n_rh,
                   cx.execute("SELECT COUNT(*) FROM dim_employe").fetchone()[0]))
VOLUMETRIE.append(("RH paie", "XLSX 2 feuilles",
                   wb["Paie mensuelle"].max_row - 1,
                   cx.execute("SELECT COUNT(*) FROM fait_paie").fetchone()[0]))

# --- Completude des champs cles apres chargement ----------------------------
COMPLETUDE = [
    ("dim_client", "code_postal", "SELECT COUNT(*) FROM dim_client WHERE code_postal IS NOT NULL"),
    ("dim_client", "courriel", "SELECT COUNT(*) FROM dim_client WHERE courriel IS NOT NULL"),
    ("dim_client", "province", "SELECT COUNT(*) FROM dim_client WHERE province IS NOT NULL"),
    ("dim_produit", "cout_standard", "SELECT COUNT(*) FROM dim_produit WHERE cout_standard IS NOT NULL"),
    ("fait_ventes", "date_paiement_id", "SELECT COUNT(*) FROM fait_ventes WHERE date_paiement_id IS NOT NULL"),
    ("fait_ventes", "canal renseigne", "SELECT COUNT(*) FROM fait_ventes WHERE canal_vente <> 'Non renseigné'"),
]
TOTAUX = {"dim_client": cx.execute("SELECT COUNT(*) FROM dim_client").fetchone()[0],
          "dim_produit": cx.execute("SELECT COUNT(*) FROM dim_produit").fetchone()[0],
          "fait_ventes": cx.execute("SELECT COUNT(*) FROM fait_ventes").fetchone()[0]}

# --- Redaction du rapport ---------------------------------------------------
n_rejetees = cx.execute("SELECT COUNT(*) FROM qualite_rejets WHERE action = 'rejetee'").fetchone()[0]
n_autres = cx.execute("SELECT COUNT(*) FROM qualite_rejets WHERE action <> 'rejetee'").fetchone()[0]

md = []
md.append("# Rapport de qualite des donnees — Boreal Distribution\n")
md.append("_Genere automatiquement par `Python/3_controle_qualite.py`. Ne pas editer a la main._\n")
md.append(f"\n**{n_rejetees} lignes rejetees** et **{n_autres} valeurs corrigees ou signalees** "
          f"sur l'ensemble des 8 sources.\n")

md.append("\n## 1. Reconciliation : anomalies injectees vs anomalies detectees\n")
md.append("Le generateur de donnees journalise chaque anomalie qu'il injecte. L'ETL journalise "
          "chaque anomalie qu'il rencontre. Les deux journaux sont produits independamment : "
          "leur confrontation mesure ce que le nettoyage laisse reellement passer.\n")
md.append("\n| Motif | Injecte | Detecte | Ecart | Regle appliquee |")
md.append("|---|---:|---:|---:|---|")
for motif, attendu, trouve, ecart, explication in lignes_reconciliation:
    marque = "0" if ecart == 0 else f"{ecart:+d}"
    md.append(f"| `{motif}` | {attendu} | {trouve} | {marque} | {explication} |")
md.append(f"| **Total** | **{sum(l[1] for l in lignes_reconciliation)}** | "
          f"**{sum(l[2] for l in lignes_reconciliation)}** | | |")

if ECARTS_ATTENDUS:
    md.append("\n### Ecarts expliques\n")
    for motif, raison in ECARTS_ATTENDUS.items():
        md.append(f"- **`{motif}`** — {raison}")

md.append("\n### Verdict\n")
if ecarts_inexpliques:
    md.append("Ecarts NON expliques (une anomalie est passee au travers du nettoyage) :\n")
    for motif, a, t in ecarts_inexpliques:
        md.append(f"- `{motif}` : {a} injectees, {t} detectees")
else:
    md.append("Aucun ecart inexplique : chaque anomalie injectee a ete retrouvee et traitee "
              "par l'ETL, ou son absence est justifiee ci-dessus.")
if non_couverts:
    md.append(f"\nAnomalies injectees sans motif correspondant : {', '.join(non_couverts)}")

md.append("\n## 2. Volumetrie par source\n")
md.append("| Source | Format | Lignes lues | Lignes chargees | Taux de retenue |")
md.append("|---|---|---:|---:|---:|")
for nom, fmt, lues, chargees in VOLUMETRIE:
    # La mise en forme francophone ne s'applique qu'aux nombres : appliquer le
    # remplacement a la ligne entiere ecraserait aussi le delimiteur ',' du format.
    md.append(f"| {nom} | {fmt} | {milliers(lues)} | {milliers(chargees)} | "
              f"{pourcent(100 * chargees / lues)} |")

md.append("\n## 3. Completude des champs cles apres chargement\n")
md.append("Aucune valeur n'a ete imputee : un champ absent reste absent. Ces taux mesurent "
          "donc ce que les systemes sources fournissent reellement.\n")
md.append("\n| Table | Champ | Renseigne | Total | Taux |")
md.append("|---|---|---:|---:|---:|")
for table, champ, requete in COMPLETUDE:
    n = cx.execute(requete).fetchone()[0]
    total = TOTAUX[table]
    md.append(f"| `{table}` | {champ} | {milliers(n)} | {milliers(total)} | "
              f"{pourcent(100 * n / total)} |")

md.append("\n## 4. Principe de traitement\n")
md.append("""
| Situation | Decision | Pourquoi |
|---|---|---|
| Format reparable (date, montant, casse, espace insecable) | **Corrigee**, ligne conservee | Le defaut est de forme, l'information metier est intacte. |
| Variante d'ecriture d'une meme valeur | **Normalisee** via table de correspondance | Une regle de casse automatique casserait `Logiciels et TI`. |
| Valeur absente | **Conservee a NULL** | Imputer une moyenne fabriquerait une donnee qui n'existe pas. |
| Cle metier introuvable (article, client) | **Rejetee**, tracee | Rattacher a un « divers » fausserait toutes les analyses par famille. |
| Ligne strictement identique | **Rejetee**, tracee | Double import : la conserver doublerait le chiffre d'affaires. |
| Deux fiches pour la meme entreprise | **Fusionnees**, ventes rattachees | Sinon le CA d'un client est eclate et la concentration sous-estimee. |
| Quantite negative sur un achat | **Conservee**, qualifiee de retour | Ce n'est pas une erreur : c'est une operation reelle. |
""".strip())

with open(SORTIE, "w", encoding="utf-8") as f:
    f.write("\n".join(md) + "\n")

print(f"Rapport ecrit : {os.path.relpath(SORTIE, BASE)}")
print(f"  {n_rejetees} lignes rejetees, {n_autres} valeurs corrigees ou signalees")
for motif, attendu, trouve, ecart, _ in lignes_reconciliation:
    etat = "OK" if ecart == 0 else ("explique" if motif in ECARTS_ATTENDUS else "A EXPLIQUER")
    print(f"  {motif:<24} injecte={attendu:>4}  detecte={trouve:>4}  {etat}")
if ecarts_inexpliques:
    print(f"\n  {len(ecarts_inexpliques)} ecart(s) inexplique(s)")
if non_couverts:
    print(f"  anomalies sans motif : {non_couverts}")
cx.close()
